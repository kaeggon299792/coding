from io import BytesIO

from openpyxl import load_workbook

from services.official_excel_import import EXPECTED_HEADERS, build_header_template


def test_header_template_has_db_sheet_and_no_data_rows():
    workbook = load_workbook(BytesIO(build_header_template()), read_only=True)
    try:
        assert workbook.sheetnames == ["DB"]
        sheet = workbook["DB"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        assert headers == EXPECTED_HEADERS
        assert sheet.max_row == 1
    finally:
        workbook.close()


def test_regular_user_can_download_import_template(monkeypatch, tmp_path):
    db_path = tmp_path / "regular-user-template.db"
    monkeypatch.setattr("config.DASHBOARD_DB_FILE", str(db_path))

    import app as app_module
    from dashboard_db import schema

    connection = schema.connect(str(db_path))
    cursor = connection.execute(
        """INSERT INTO dashboard_users
           (username, password_hash, role, is_active, created_at)
           VALUES ('regular', 'unused', 'user', 1, '2026-07-29T00:00:00')"""
    )
    user_id = cursor.lastrowid
    connection.execute(
        """INSERT INTO dashboard_user_permissions
           (user_id, permission_code, allowed, updated_at)
           VALUES (?, 'official_docs', 1, '2026-07-29T00:00:00')""",
        (user_id,),
    )
    connection.commit()
    connection.close()

    app_module.app.config["TESTING"] = True
    with app_module.app.test_client() as client:
        with client.session_transaction() as browser_session:
            browser_session["user_id"] = user_id
            browser_session["username"] = "regular"
            browser_session["role"] = "user"
        response = client.get("/official-docs/import-excel/template")
        filtered_export = client.get(
            "/official-docs/export?export_type=filtered"
        )

    assert response.status_code == 200
    assert response.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert filtered_export.status_code == 200
    assert filtered_export.mimetype == response.mimetype
    workbook = load_workbook(BytesIO(response.data), read_only=True)
    try:
        assert [cell.value for cell in workbook["DB"][1]] == EXPECTED_HEADERS
    finally:
        workbook.close()


def test_import_template_download_enforces_database_approval_and_permission(
    monkeypatch, tmp_path
):
    db_path = tmp_path / "template-download-permissions.db"
    monkeypatch.setattr("config.DASHBOARD_DB_FILE", str(db_path))

    import app as app_module
    from dashboard_db import schema

    connection = schema.connect(str(db_path))

    def add_user(username, role, approval_status):
        cursor = connection.execute(
            """INSERT INTO dashboard_users
               (username, password_hash, role, is_active, approval_status, created_at)
               VALUES (?, 'unused', ?, 1, ?, '2026-08-04T00:00:00')""",
            (username, role, approval_status),
        )
        return cursor.lastrowid

    admin_id = add_user("template-admin", "admin", "approved")
    pending_id = add_user("template-pending", "user", "pending")
    denied_id = add_user("template-denied", "user", "approved")
    connection.executemany(
        """INSERT INTO dashboard_user_permissions
           (user_id, permission_code, allowed, updated_at)
           VALUES (?, 'official_docs', ?, '2026-08-04T00:00:00')""",
        ((pending_id, 1), (denied_id, 0)),
    )
    connection.commit()
    connection.close()

    app_module.app.config["TESTING"] = True
    with app_module.app.test_client() as client:
        anonymous = client.get("/official-docs/import-excel/template")

        with client.session_transaction() as browser_session:
            browser_session["user_id"] = admin_id
            browser_session["username"] = "template-admin"
            browser_session["role"] = "admin"
        admin = client.get("/official-docs/import-excel/template")

        with client.session_transaction() as browser_session:
            browser_session.clear()
            browser_session["user_id"] = pending_id
            browser_session["username"] = "template-pending"
            # A forged client-side role must not override the database row.
            browser_session["role"] = "admin"
        pending = client.get(
            "/official-docs/import-excel/template?approval_status=approved"
        )

        with client.session_transaction() as browser_session:
            browser_session.clear()
            browser_session["user_id"] = denied_id
            browser_session["username"] = "template-denied"
            browser_session["role"] = "user"
        denied = client.get("/official-docs/import-excel/template")

    assert anonymous.status_code == 302
    assert "/login" in anonymous.headers["Location"]
    assert admin.status_code == 200
    assert pending.status_code == 403
    assert denied.status_code == 403
