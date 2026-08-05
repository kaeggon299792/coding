"""Database-backed localization inventory, scan, QA and bulk interchange."""

from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import re
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from utils import now_kst


HANGUL_RE = re.compile(r"[가-힣]")
TAG_RE = re.compile(r"<[^>]+>")
HTML_TEXT_RE = re.compile(r">\s*([^<>{%]*[가-힣][^<>{%]*)\s*<")
ATTRIBUTE_RE = re.compile(
    r"(?:placeholder|title|aria-label|aria-description|data-confirm)\s*=\s*['\"]([^'\"]*[가-힣][^'\"]*)['\"]",
    re.IGNORECASE,
)
QUOTED_RE = re.compile(r"(?P<q>['\"])(?P<text>[^'\"\r\n]*[가-힣][^'\"\r\n]*)(?P=q)")
VARIABLE_RE = re.compile(r"\{\{?\s*[\w.]+\s*\}?\}|%\([^)]+\)[a-zA-Z]|%[a-zA-Z]")
MARKDOWN_RE = re.compile(r"(?:\[[^\]]+\]\([^)]+\)|[*_`#]{1,3})")
ALLOWED_STATUS = {"Pending", "Completed", "Ignored"}
ALLOWED_PRIORITY = {"Critical", "High", "Medium", "Low"}
PRIORITY_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
TRANSLATION_STYLES = {
    "literal": "원문의 의미와 문장 구조를 가능한 한 충실하게 번역합니다.",
    "natural": "대상 언어 사용자가 자연스럽게 이해하도록 번역합니다.",
    "casino": "카지노·복합리조트 산업에서 통용되는 전문 용어를 우선합니다.",
    "business": "간결하고 전문적인 비즈니스 문체를 사용합니다.",
    "marketing": "정확성을 유지하면서 설득력 있는 마케팅 문체를 사용합니다.",
}
PROMPT_LANGUAGES = {
    "en": {"target_name": "영어", "output_label": "EN"},
    "ja": {"target_name": "일본어", "output_label": "JA"},
    "yue-HK": {"target_name": "광둥어(홍콩 번체)", "output_label": "YUE"},
}
PROMPT_MAX_ITEMS = 150
PROMPT_MAX_CHARS = 60000
ADMIN_ONLY_TEMPLATES = {
    "templates/admin_action_items.html",
    "templates/ai_settings.html",
    "templates/admin_logs.html",
    "templates/localization_admin.html",
    "templates/localization_prompt.html",
    "templates/localization_work.html",
    "templates/localization_qa.html",
    "templates/user_management.html",
    "templates/official_docs/settings.html",
}
IGNORED_I18N_BLOCK_RE = re.compile(
    r"<(?P<tag>[a-z][\w:-]*)\b[^>]*\bdata-i18n-ignore\b[^>]*>"
    r".*?</(?P=tag)\s*>",
    re.IGNORECASE | re.DOTALL,
)


def _timestamp():
    return now_kst().isoformat(timespec="microseconds")


def _hash(text):
    return hashlib.sha256(text.strip().encode("utf-8")).hexdigest()


def _clean(text):
    return html.unescape(str(text or "")).strip()


def _priority(page, component, string_type):
    haystack = f"{page} {component} {string_type}".lower()
    if any(word in haystack for word in ("menu", "nav", "login", "register", "payment", "booking")):
        return "Critical"
    if any(word in haystack for word in ("company", "news", "research", "ai opinion", "기업", "뉴스", "리서치")):
        return "High"
    if any(word in haystack for word in ("tip", "notice", "faq", "공지", "자료")):
        return "Medium"
    if "admin" in haystack or "관리자" in haystack:
        return "Low"
    return "Medium"


def _key(text, prefix="UI"):
    return f"{prefix}_{_hash(text)[:16].upper()}"


def register_string(connection, source_text, *, page="Unknown", component="Content",
                    string_type="UI", source_kind="runtime", source_path="", locator="",
                    language_key=None):
    """Register or refresh one deduplicated source string and its reference."""
    text = _clean(source_text)
    if not text or not HANGUL_RE.search(text) or len(text) > 20000:
        return None
    now = _timestamp()
    digest = _hash(text)
    row = connection.execute(
        "SELECT id, source_text FROM localization_strings WHERE source_hash=?", (digest,)
    ).fetchone()
    created = row is None
    if created:
        cursor = connection.execute(
            """INSERT INTO localization_strings
                   (language_key, source_text, source_hash, page_name, component,
                    string_type, priority, status, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, 'Pending', ?, ?)""",
            (language_key or _key(text), text, digest, page, component, string_type,
             _priority(page, component, string_type), now, now),
        )
        string_id = cursor.lastrowid
        connection.execute(
            """INSERT INTO localization_events
                   (string_id, event_type, detail, created_at)
               VALUES (?, 'pending_created', ?, ?)""",
            (string_id, source_path[:500], now),
        )
    else:
        string_id = row["id"]
        connection.execute(
            "UPDATE localization_strings SET deleted_at=NULL WHERE id=?", (string_id,)
        )
    if source_path:
        connection.execute(
            """INSERT INTO localization_references
                   (string_id, source_kind, source_path, locator, last_seen_at)
               VALUES (?, ?, ?, ?, ?)
               ON CONFLICT(string_id, source_kind, source_path, locator)
               DO UPDATE SET last_seen_at=excluded.last_seen_at""",
            (string_id, source_kind, source_path[:500], locator[:200], now),
        )
    return string_id


def register_rendered_strings(connection, strings, *, page, source_path):
    """Register Korean copy observed on a public rendered page in one batch."""

    registered = 0
    for index, text in enumerate(strings):
        if register_string(
            connection,
            text,
            page=page,
            component="Rendered Content",
            string_type="UI",
            source_kind="rendered",
            source_path=source_path,
            locator=f"text:{index}",
        ):
            registered += 1
    connection.commit()
    return registered


def _dynamic_values(value):
    """Yield strings from scalar or JSON-backed public content fields."""

    if value is None:
        return
    parsed = value
    if isinstance(value, str):
        stripped = value.strip()
        if stripped[:1] in ("[", "{"):
            try:
                parsed = json.loads(stripped)
            except (TypeError, ValueError):
                parsed = value
    if isinstance(parsed, dict):
        for item in parsed.values():
            yield from _dynamic_values(item)
    elif isinstance(parsed, (list, tuple)):
        for item in parsed:
            yield from _dynamic_values(item)
    elif isinstance(parsed, str):
        yield parsed


def _scan_file(path, root):
    try:
        source = path.read_text(encoding="utf-8")
    except (OSError, UnicodeError):
        return []
    relative = path.relative_to(root).as_posix()
    if relative in ADMIN_ONLY_TEMPLATES:
        return []
    source = IGNORED_I18N_BLOCK_RE.sub("", source)
    page = path.stem.replace("_", " ").title()
    matches = []
    patterns = (HTML_TEXT_RE, ATTRIBUTE_RE) if path.suffix == ".html" else (QUOTED_RE,)
    for pattern in patterns:
        for match in pattern.finditer(source):
            text = _clean(match.group(1) if pattern is not QUOTED_RE else match.group("text"))
            text = re.sub(r"\s+", " ", text)
            if not text or len(text) > 2000 or not HANGUL_RE.search(text):
                continue
            line = source.count("\n", 0, match.start()) + 1
            component = "Template" if path.suffix == ".html" else "Server Message"
            string_type = "UI" if path.suffix == ".html" else "Message"
            matches.append((text, page, component, string_type, relative, f"line:{line}"))
    return matches


def scan_project(connection, project_root):
    """Inventory static source plus supported dynamic content; never modifies source files."""
    root = Path(project_root).resolve()
    scanned = registered = 0
    scan_started = _timestamp()
    for folder, suffixes in (
        ("templates", {".html"}), ("static/js", {".js"}),
        ("auth", {".py"}), ("services", {".py"}),
    ):
        base = root / folder
        if not base.exists():
            continue
        for path in base.rglob("*"):
            if not path.is_file() or path.suffix not in suffixes:
                continue
            scanned += 1
            for text, page, component, kind, relative, locator in _scan_file(path, root):
                if register_string(connection, text, page=page, component=component,
                                   string_type=kind, source_kind="file",
                                   source_path=relative, locator=locator):
                    registered += 1
    for filename in ("app.py", "official_docs.py", "tips.py", "localization.py"):
        path = root / filename
        if not path.is_file():
            continue
        scanned += 1
        for text, page, component, kind, relative, locator in _scan_file(path, root):
            if register_string(connection, text, page=page, component=component,
                               string_type=kind, source_kind="file",
                               source_path=relative, locator=locator):
                registered += 1

    catalog_path = root / "translations" / "catalog.json"
    if catalog_path.is_file():
        catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
        for source_text, translated in catalog.get("text", {}).items():
            string_id = register_string(
                connection, source_text, page="Global", component="Catalog",
                string_type="UI", source_kind="catalog",
                source_path="translations/catalog.json", language_key=_key(source_text, "CATALOG"),
            )
            if string_id and translated:
                save_translation(connection, string_id, "en", translated, actor_id=None)

    # Dynamic sources are registered without translating or changing their data.
    dynamic = (
        ("company", "monitored_companies", ("name",), "기업정보", "Company", ""),
        ("company_profile", "company_research_profiles",
         ("company_name", "legal_name", "ceo_names", "headquarters", "business_summary",
          "strategy_summary", "key_assets_json", "opportunities_json", "risks_json",
          "financials_json", "sources_json"), "기업정보", "Company", ""),
        ("law", "monitored_laws", ("law_name", "notes"), "법령·규제", "Law", "active=1"),
        ("law_analysis", "law_analysis",
         ("ai_summary", "affected_scope", "company_impact", "action_needed"),
         "법령·규제", "Law Analysis", ""),
        ("assembly_bill", "legislative_bills",
         ("bill_kind", "bill_name", "proposer_kind", "proposer_name", "committee_name",
          "committee_result", "plenary_result", "process_stage", "pass_status",
          "matched_keyword", "ai_summary", "impact_direction", "impact_level",
          "impact_reason", "action_needed", "analysis_error"),
         "입법동향", "Legislation", ""),
        ("government_notice", "government_legislative_notices",
         ("notice_name", "law_type", "ministry_name", "attachment_name",
          "announcement_type", "matched_keyword"), "입법동향", "Legislation", ""),
        ("notice", "community_posts", ("title", "content"), "공지", "Notice", "board_type='notice'"),
        ("board", "community_posts", ("title", "content"), "자유 게시판", "Content", "board_type='community'"),
        ("research", "research_documents",
         ("title", "ai_summary", "industry_impact", "investment_stance"), "리서치", "Research", ""),
        ("tip", "tips_articles", ("title", "summary", "body"), "자료실", "Tips", "is_deleted=0"),
        ("casino_glossary", "casino_glossary_terms",
         ("term_ko", "term_en", "definition", "easy_explanation", "aliases"),
         "카지노 용어집", "Glossary", "is_deleted=0 AND is_public=1"),
    )
    tables = {row[0] for row in connection.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()}
    for kind, table, fields, page, string_type, where in dynamic:
        if table not in tables:
            continue
        columns = {row[1] for row in connection.execute(f"PRAGMA table_info({table})")}
        fields = tuple(field for field in fields if field in columns)
        if not fields:
            continue
        sql = f"SELECT id, {', '.join(fields)} FROM {table}"
        if where:
            sql += f" WHERE {where}"
        for row in connection.execute(sql).fetchall():
            for field in fields:
                for value_index, value in enumerate(_dynamic_values(row[field])):
                    if register_string(connection, value, page=page, component="Dynamic Content",
                                       string_type=string_type, source_kind="database",
                                       source_path=f"{table}:{row['id']}",
                                       locator=f"{field}:{value_index}"):
                        registered += 1

    # News lives in the separate read-only news_history.db. Inventory every
    # public title, category, issue title and AI summary without modifying it.
    try:
        from services import news_reader
        news_rows = news_reader.localization_content_rows()
    except Exception:
        news_rows = []
    for row in news_rows:
        if register_string(
            connection,
            row.get("source_text"),
            page="국내 뉴스",
            component=row.get("component") or "News",
            string_type="News",
            source_kind="external_database",
            source_path=f"news_history:{row.get('source_id')}",
            locator=row.get("field") or "content",
        ):
            registered += 1

    # A full scan owns file/catalog references. References not seen in this run
    # represent removed or modified source strings and are retired non-destructively.
    connection.execute(
        "DELETE FROM localization_references WHERE source_kind IN ('file','catalog','external_database') AND last_seen_at<?",
        (scan_started,),
    )
    stale = connection.execute(
        """SELECT s.id FROM localization_strings s
           WHERE s.deleted_at IS NULL AND NOT EXISTS
             (SELECT 1 FROM localization_references r WHERE r.string_id=s.id)"""
    ).fetchall()
    for row in stale:
        connection.execute(
            "UPDATE localization_strings SET deleted_at=?, updated_at=? WHERE id=?",
            (scan_started, scan_started, row["id"]),
        )
        connection.execute(
            "INSERT INTO localization_events (string_id,event_type,created_at) VALUES (?,'deleted',?)",
            (row["id"], scan_started),
        )
    connection.commit()
    return {"files_scanned": scanned, "strings_seen": registered}


def scan_if_due(connection, project_root, interval_minutes=60):
    """Run at most once per interval when an administrator renders the site."""
    row = connection.execute(
        "SELECT setting_value FROM site_settings WHERE setting_key='localization_last_scan_at'"
    ).fetchone()
    if row:
        try:
            last_scan = datetime.fromisoformat(row["setting_value"])
            if now_kst() - last_scan < timedelta(minutes=interval_minutes):
                return None
        except (TypeError, ValueError):
            pass
    result = scan_project(connection, project_root)
    now = _timestamp()
    connection.execute(
        """INSERT INTO site_settings (setting_key, setting_value, updated_by, updated_at)
           VALUES ('localization_last_scan_at', ?, NULL, ?)
           ON CONFLICT(setting_key) DO UPDATE SET
             setting_value=excluded.setting_value, updated_at=excluded.updated_at""",
        (now, now),
    )
    connection.commit()
    return result


def save_translation(connection, string_id, language_code, translated_text, actor_id=None,
                     *, status="Completed", record_event=True):
    if status not in ALLOWED_STATUS:
        raise ValueError("invalid status")
    language = connection.execute(
        "SELECT 1 FROM localization_languages WHERE language_code=? AND is_active=1",
        (language_code,),
    ).fetchone()
    if not language or language_code == "ko":
        raise ValueError("invalid language")
    text = str(translated_text or "").strip()
    if status == "Completed" and not text:
        raise ValueError("completed translation cannot be empty")
    now = _timestamp()
    connection.execute(
        """INSERT INTO localization_translations
               (string_id, language_code, translated_text, status, translated_by,
                created_at, updated_at, last_translated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(string_id, language_code) DO UPDATE SET
               translated_text=excluded.translated_text, status=excluded.status,
               translated_by=excluded.translated_by, updated_at=excluded.updated_at,
               last_translated_at=excluded.last_translated_at""",
        (string_id, language_code, text or None, status, actor_id, now, now,
         now if status == "Completed" else None),
    )
    connection.execute(
        "UPDATE localization_strings SET updated_at=? WHERE id=?", (now, string_id)
    )
    if record_event:
        connection.execute(
            """INSERT INTO localization_events
                   (string_id, event_type, language_code, actor_id, created_at)
               VALUES (?, ?, ?, ?, ?)""",
            (string_id, "translation_completed" if status == "Completed" else "status_changed",
             language_code, actor_id, now),
        )


def dashboard_summary(connection, language_code="en"):
    counts = {row["status"]: row["count"] for row in connection.execute(
        """SELECT CASE WHEN s.status='Ignored' THEN 'Ignored'
                       ELSE COALESCE(t.status, 'Pending') END AS status, COUNT(*) AS count
           FROM localization_strings s
           LEFT JOIN localization_translations t
             ON t.string_id=s.id AND t.language_code=?
           WHERE s.deleted_at IS NULL GROUP BY CASE WHEN s.status='Ignored' THEN 'Ignored'
                       ELSE COALESCE(t.status, 'Pending') END""",
        (language_code,),
    ).fetchall()}
    total = sum(counts.values())
    completed = counts.get("Completed", 0)
    coverage = round(completed * 100 / total, 2) if total else 100.0
    health = "Excellent" if coverage >= 98 else "Good" if coverage >= 90 else "Needs Review"
    since = (now_kst() - timedelta(hours=24)).isoformat(timespec="seconds")
    recent_counts = {row["event_type"]: row["count"] for row in connection.execute(
        "SELECT event_type, COUNT(*) AS count FROM localization_events WHERE created_at>=? GROUP BY event_type",
        (since,),
    ).fetchall()}
    events = connection.execute(
        """SELECT e.*, s.language_key, s.source_text FROM localization_events e
           LEFT JOIN localization_strings s ON s.id=e.string_id
           WHERE e.created_at>=? ORDER BY e.created_at DESC LIMIT 30""", (since,)
    ).fetchall()
    return {"counts": counts, "total": total, "completed": completed,
            "pending": counts.get("Pending", 0), "ignored": counts.get("Ignored", 0),
            "coverage": coverage, "health": health,
            "recent_counts": recent_counts, "events": [dict(row) for row in events]}


def dashboard_summary_all(connection):
    """Summarize every active target-language/string pair."""
    counts = {row["status"]: row["count"] for row in connection.execute(
        """SELECT CASE WHEN s.status='Ignored' THEN 'Ignored'
                       ELSE COALESCE(t.status, 'Pending') END AS status,
                  COUNT(*) AS count
           FROM localization_strings s
           JOIN localization_languages l ON l.is_active=1 AND l.is_source=0
           LEFT JOIN localization_translations t
             ON t.string_id=s.id AND t.language_code=l.language_code
           WHERE s.deleted_at IS NULL
           GROUP BY CASE WHEN s.status='Ignored' THEN 'Ignored'
                         ELSE COALESCE(t.status, 'Pending') END"""
    ).fetchall()}
    total = sum(counts.values())
    completed = counts.get("Completed", 0)
    coverage = round(completed * 100 / total, 2) if total else 100.0
    health = "Excellent" if coverage >= 98 else "Good" if coverage >= 90 else "Needs Review"
    since = (now_kst() - timedelta(hours=24)).isoformat(timespec="seconds")
    recent_counts = {row["event_type"]: row["count"] for row in connection.execute(
        "SELECT event_type, COUNT(*) AS count FROM localization_events WHERE created_at>=? GROUP BY event_type",
        (since,),
    ).fetchall()}
    events = connection.execute(
        """SELECT e.*, s.language_key, s.source_text FROM localization_events e
           LEFT JOIN localization_strings s ON s.id=e.string_id
           WHERE e.created_at>=? ORDER BY e.created_at DESC LIMIT 30""",
        (since,),
    ).fetchall()
    return {"counts": counts, "total": total, "completed": completed,
            "pending": counts.get("Pending", 0), "ignored": counts.get("Ignored", 0),
            "coverage": coverage, "health": health,
            "recent_counts": recent_counts, "events": [dict(row) for row in events]}


def list_strings(connection, *, language_code="en", query="", status="", priority="",
                 sort="newest", page=1, per_page=50):
    clauses = ["s.deleted_at IS NULL"]
    params = [language_code]
    if query:
        clauses.append("(s.source_text LIKE ? OR s.language_key LIKE ? OR s.page_name LIKE ? OR s.string_type LIKE ?)")
        needle = f"%{query[:200]}%"
        params.extend([needle] * 4)
    if status in ALLOWED_STATUS:
        clauses.append("(CASE WHEN s.status='Ignored' THEN 'Ignored' ELSE COALESCE(t.status, 'Pending') END)=?")
        params.append(status)
    if priority in ALLOWED_PRIORITY:
        clauses.append("s.priority=?")
        params.append(priority)
    order = {
        "oldest": "s.created_at ASC", "page": "s.page_name, s.language_key",
        "type": "s.string_type, s.language_key",
        "priority": "CASE s.priority WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END, s.updated_at DESC",
    }.get(sort, "s.updated_at DESC")
    where = " AND ".join(clauses)
    total = connection.execute(
        f"""SELECT COUNT(*) FROM localization_strings s
            LEFT JOIN localization_translations t ON t.string_id=s.id AND t.language_code=?
            WHERE {where}""", params,
    ).fetchone()[0]
    offset = (max(1, page) - 1) * per_page
    rows = connection.execute(
        f"""SELECT s.*, t.translated_text,
                   CASE WHEN s.status='Ignored' THEN 'Ignored' ELSE COALESCE(t.status, 'Pending') END AS effective_status,
                   t.last_translated_at,
                   (SELECT COUNT(*) FROM localization_references r WHERE r.string_id=s.id) AS reference_count
            FROM localization_strings s
            LEFT JOIN localization_translations t ON t.string_id=s.id AND t.language_code=?
            WHERE {where} ORDER BY {order} LIMIT ? OFFSET ?""",
        (*params, per_page, offset),
    ).fetchall()
    return [dict(row) for row in rows], total


def list_strings_all(connection, *, query="", status="", priority="",
                     sort="newest", page=1, per_page=50):
    """List one row per active target language and source string."""
    clauses = ["s.deleted_at IS NULL", "l.is_active=1", "l.is_source=0"]
    params = []
    if query:
        clauses.append("(s.source_text LIKE ? OR s.language_key LIKE ? OR s.page_name LIKE ? OR s.string_type LIKE ? OR l.display_name LIKE ?)")
        needle = f"%{query[:200]}%"
        params.extend([needle] * 5)
    effective_status = "CASE WHEN s.status='Ignored' THEN 'Ignored' ELSE COALESCE(t.status, 'Pending') END"
    if status in ALLOWED_STATUS:
        clauses.append(f"({effective_status})=?")
        params.append(status)
    if priority in ALLOWED_PRIORITY:
        clauses.append("s.priority=?")
        params.append(priority)
    order = {
        "oldest": "s.created_at ASC, l.language_code",
        "page": "s.page_name, s.language_key, l.language_code",
        "type": "s.string_type, s.language_key, l.language_code",
        "priority": "CASE s.priority WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END, s.updated_at DESC, l.language_code",
    }.get(sort, "s.updated_at DESC, l.language_code")
    where = " AND ".join(clauses)
    joins = """JOIN localization_languages l
                 LEFT JOIN localization_translations t
                   ON t.string_id=s.id AND t.language_code=l.language_code"""
    total = connection.execute(
        f"SELECT COUNT(*) FROM localization_strings s {joins} WHERE {where}",
        params,
    ).fetchone()[0]
    offset = (max(1, page) - 1) * per_page
    rows = connection.execute(
        f"""SELECT s.*, t.translated_text, ({effective_status}) AS effective_status,
                   t.last_translated_at,
                   l.language_code AS target_language_code,
                   l.display_name AS target_language_name,
                   (SELECT COUNT(*) FROM localization_references r WHERE r.string_id=s.id) AS reference_count
            FROM localization_strings s {joins}
            WHERE {where} ORDER BY {order} LIMIT ? OFFSET ?""",
        (*params, per_page, offset),
    ).fetchall()
    return [dict(row) for row in rows], total


def references(connection, string_id):
    return [dict(row) for row in connection.execute(
        "SELECT * FROM localization_references WHERE string_id=? ORDER BY source_kind, source_path",
        (string_id,),
    ).fetchall()]


def list_glossary(connection, language_code="en", *, active_only=False):
    sql = "SELECT * FROM localization_glossary WHERE language_code=?"
    if active_only:
        sql += " AND is_active=1"
    sql += " ORDER BY source_text COLLATE NOCASE, id"
    return [dict(row) for row in connection.execute(sql, (language_code,)).fetchall()]


def save_glossary(connection, source_text, target_text, language_code="en",
                  actor_id=None, glossary_id=None):
    source = _clean(source_text)
    target = _clean(target_text)
    if not source or len(source) > 300 or not target or len(target) > 500:
        raise ValueError("용어는 원문 300자, 번역 500자 이내로 입력해주세요.")
    if not connection.execute(
        "SELECT 1 FROM localization_languages WHERE language_code=? AND is_active=1",
        (language_code,),
    ).fetchone():
        raise ValueError("대상 언어를 확인해주세요.")
    duplicate = connection.execute(
        """SELECT id FROM localization_glossary
           WHERE language_code=? AND source_text=? AND (? IS NULL OR id<>?)""",
        (language_code, source, glossary_id, glossary_id),
    ).fetchone()
    if duplicate:
        raise ValueError("같은 언어에 이미 등록된 용어입니다.")
    now = _timestamp()
    if glossary_id:
        cursor = connection.execute(
            """UPDATE localization_glossary
               SET source_text=?, target_text=?, is_active=1, updated_at=?, updated_by=?
               WHERE id=? AND language_code=?""",
            (source, target, now, actor_id, glossary_id, language_code),
        )
        if not cursor.rowcount:
            raise ValueError("용어집 항목을 찾을 수 없습니다.")
    else:
        connection.execute(
            """INSERT INTO localization_glossary
                   (source_text,target_text,language_code,is_active,created_at,updated_at,updated_by)
               VALUES (?,?,?,1,?,?,?)""",
            (source, target, language_code, now, now, actor_id),
        )


def deactivate_glossary(connection, glossary_id, language_code="en"):
    cursor = connection.execute(
        "UPDATE localization_glossary SET is_active=0, updated_at=? WHERE id=? AND language_code=?",
        (_timestamp(), glossary_id, language_code),
    )
    if not cursor.rowcount:
        raise ValueError("용어집 항목을 찾을 수 없습니다.")


def _prompt_language(connection, language_code):
    row = connection.execute(
        "SELECT display_name FROM localization_languages WHERE language_code=? AND is_active=1",
        (language_code,),
    ).fetchone()
    if not row or language_code == "ko":
        raise ValueError("지원하지 않는 번역 대상 언어입니다.")
    configured = PROMPT_LANGUAGES.get(language_code)
    if configured:
        return configured
    label = re.sub(r"[^A-Z0-9]+", "_", language_code.upper()).strip("_")
    return {"target_name": row["display_name"], "output_label": label}


def _prompt_entry(row, output_label):
    return (
        "--------------------------------\n\n"
        f"ID={row['language_key']}\n\n"
        f"TYPE={row['string_type']}\n"
        f"PAGE={row['page_name']}\n\n"
        "KOREAN:\n"
        f"{row['source_text']}\n\n"
        f"{output_label}:\n\n"
    )


def _prompt_instructions(style, glossary, target_name, output_label):
    style_rule = TRANSLATION_STYLES.get(style, TRANSLATION_STYLES["natural"])
    glossary_text = "\n".join(
        f"- {item['source_text']} → {item['target_text']}" for item in glossary
    ) or "- 등록된 용어 없음"
    return f"""당신은 카지노 산업 전문 번역가입니다.

다음 한국어 문자열을 자연스러운 {target_name}로 번역하세요.

번역 규칙

1. 카지노 업계에서 사용하는 용어를 사용합니다.
2. 브랜드명, 회사명, 카지노명은 임의로 번역하지 않습니다.
3. HTML 태그를 수정하지 않습니다.
4. Markdown 문법을 수정하지 않습니다.
5. Placeholder(예: {{name}}, {{{{count}}}}, %s)는 그대로 유지합니다.
6. 줄바꿈을 유지합니다.
7. 번역되지 않은 항목만 작성합니다.
8. 설명은 작성하지 않습니다.
9. ID와 {output_label} 출력 형식을 반드시 유지합니다.
10. 전체 번역 결과를 하나의 ```text 코드블록 안에 작성합니다.
11. 코드블록 밖에는 어떤 설명, 주석 또는 인사말도 작성하지 않습니다.
12. 선택한 번역 스타일: {style_rule}

프로젝트 공통 용어집
{glossary_text}

아래 내용을 번역하세요.

=========================

"""


def generate_translation_chunks(connection, language_code, string_ids, *,
                                mode="prompt", style="natural",
                                max_items=PROMPT_MAX_ITEMS,
                                max_chars=PROMPT_MAX_CHARS):
    prompt_language = _prompt_language(connection, language_code)
    output_label = prompt_language["output_label"]
    ids = []
    for value in string_ids:
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            continue
        if parsed > 0 and parsed not in ids:
            ids.append(parsed)
    if not ids:
        raise ValueError("번역할 항목을 하나 이상 선택해주세요.")
    if len(ids) > 500:
        raise ValueError("한 번에 최대 500개 항목을 선택할 수 있습니다.")
    max_items = max(1, min(int(max_items or PROMPT_MAX_ITEMS), 500))
    max_chars = max(3000, min(int(max_chars or PROMPT_MAX_CHARS), 100000))
    placeholders = ",".join("?" for _ in ids)
    rows = [dict(row) for row in connection.execute(
        f"""SELECT s.id,s.language_key,s.source_text,s.page_name,s.string_type
            FROM localization_strings s
            LEFT JOIN localization_translations t
              ON t.string_id=s.id AND t.language_code=?
            WHERE s.id IN ({placeholders}) AND s.deleted_at IS NULL
              AND s.status<>'Ignored' AND COALESCE(t.status,'Pending')<>'Completed'
            ORDER BY CASE s.priority WHEN 'Critical' THEN 0 WHEN 'High' THEN 1
                     WHEN 'Medium' THEN 2 ELSE 3 END, s.id""",
        (language_code, *ids),
    ).fetchall()]
    if not rows:
        raise ValueError("선택 항목 중 미번역 대상이 없습니다.")
    entries = [_prompt_entry(row, output_label) for row in rows]
    groups, current, current_size = [], [], 0
    for entry in entries:
        if current and (len(current) >= max_items or current_size + len(entry) > max_chars):
            groups.append(current); current, current_size = [], 0
        current.append(entry); current_size += len(entry)
    if current:
        groups.append(current)
    glossary = list_glossary(connection, language_code, active_only=True)
    prefix = _prompt_instructions(
        style, glossary, prompt_language["target_name"], output_label
    ) if mode == "prompt" else ""
    suffix = (
        "=========================\n\n"
        "위 규칙을 모두 준수하고, 출력 형식은 절대 변경하지 마십시오. "
        "설명이나 주석은 작성하지 말고, 전체 번역 결과만 하나의 ```text 코드블록으로 반환하십시오."
        if mode == "prompt" else ""
    )
    return [prefix + "".join(group) + suffix for group in groups]


def _active_prompt_languages(connection):
    languages = []
    for row in connection.execute(
        """SELECT language_code, display_name FROM localization_languages
           WHERE is_active=1 AND is_source=0 ORDER BY language_code"""
    ).fetchall():
        configured = _prompt_language(connection, row["language_code"])
        languages.append({
            "language_code": row["language_code"],
            "display_name": row["display_name"],
            **configured,
        })
    if not languages:
        raise ValueError("활성화된 번역 대상 언어가 없습니다.")
    return languages


def generate_all_translation_chunks(connection, string_ids, *, mode="prompt",
                                    style="natural", max_items=PROMPT_MAX_ITEMS,
                                    max_chars=PROMPT_MAX_CHARS):
    """Generate one clipboard payload containing every active target language."""
    languages = _active_prompt_languages(connection)
    ids = []
    for value in string_ids:
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            continue
        if parsed > 0 and parsed not in ids:
            ids.append(parsed)
    if not ids:
        raise ValueError("번역할 항목을 하나 이상 선택해주세요.")
    if len(ids) > 500:
        raise ValueError("한 번에 최대 500개 항목을 선택할 수 있습니다.")

    placeholders = ",".join("?" for _ in ids)
    rows = [dict(row) for row in connection.execute(
        f"""SELECT id, language_key, source_text, page_name, string_type
            FROM localization_strings
            WHERE id IN ({placeholders}) AND deleted_at IS NULL
              AND status<>'Ignored'
            ORDER BY CASE priority WHEN 'Critical' THEN 0 WHEN 'High' THEN 1
                     WHEN 'Medium' THEN 2 ELSE 3 END, id""",
        ids,
    ).fetchall()]
    completed = {
        (row["string_id"], row["language_code"])
        for row in connection.execute(
            f"""SELECT string_id, language_code FROM localization_translations
                WHERE string_id IN ({placeholders}) AND status='Completed'""",
            ids,
        ).fetchall()
    }
    entries = []
    for row in rows:
        pending_languages = [
            language for language in languages
            if (row["id"], language["language_code"]) not in completed
        ]
        if not pending_languages:
            continue
        fields = "\n".join(
            f"{language['output_label']}:\n" for language in pending_languages
        )
        entries.append(
            "--------------------------------\n\n"
            f"ID={row['language_key']}\n\n"
            f"TYPE={row['string_type']}\n"
            f"PAGE={row['page_name']}\n\n"
            f"KOREAN:\n{row['source_text']}\n\n{fields}\n"
        )
    if not entries:
        raise ValueError("선택 항목은 모든 언어의 번역이 완료됐습니다.")

    max_items = max(1, min(int(max_items or PROMPT_MAX_ITEMS), 500))
    max_chars = max(3000, min(int(max_chars or PROMPT_MAX_CHARS), 100000))
    groups, current, current_size = [], [], 0
    for entry in entries:
        if current and (len(current) >= max_items or current_size + len(entry) > max_chars):
            groups.append(current)
            current, current_size = [], 0
        current.append(entry)
        current_size += len(entry)
    if current:
        groups.append(current)

    if mode != "prompt":
        return ["".join(group) for group in groups]
    language_lines = "\n".join(
        f"- {language['output_label']}: {language['display_name']}"
        for language in languages
    )
    glossary_lines = []
    for language in languages:
        terms = list_glossary(
            connection, language["language_code"], active_only=True
        )
        if terms:
            glossary_lines.append(
                f"[{language['output_label']}]\n" + "\n".join(
                    f"- {term['source_text']} → {term['target_text']}" for term in terms
                )
            )
    prefix = f"""당신은 카지노 산업 전문 번역가입니다.

각 KOREAN 문장을 아래의 모든 대상 언어로 번역하세요. 이미 완료된 언어는 입력칸이 없으며, 표시된 언어 라벨만 채웁니다.

대상 언어
{language_lines}

규칙
1. ID, TYPE, PAGE, KOREAN과 언어 라벨을 변경하지 않습니다.
2. HTML, Markdown, 줄바꿈, 변수와 placeholder를 보존합니다.
3. 브랜드명과 회사명은 임의로 바꾸지 않습니다.
4. 설명이나 주석 없이 전체 결과를 하나의 ```text 코드블록으로 반환합니다.
5. 번역 스타일은 {TRANSLATION_STYLES.get(style, TRANSLATION_STYLES['natural'])}

프로젝트 공통 용어집
{chr(10).join(glossary_lines) or '- 등록된 용어 없음'}

=========================\n\n"""
    suffix = (
        "=========================\n\n"
        "모든 ID와 표시된 언어 라벨을 유지하고 번역 결과만 채워주세요."
    )
    return [prefix + "".join(group) + suffix for group in groups]


def import_ai_translation_text_all(connection, payload, actor_id=None):
    """Import a combined AI response containing multiple language labels."""
    text = str(payload or "")
    if not text.strip():
        raise ValueError("AI 번역 결과를 붙여 넣어주세요.")
    if len(text) > 2_000_000:
        raise ValueError("붙여넣기 내용은 2MB 이하여야 합니다.")
    languages = _active_prompt_languages(connection)
    by_label = {item["output_label"]: item["language_code"] for item in languages}
    label_pattern = "|".join(re.escape(label) for label in by_label)
    block_pattern = re.compile(
        r"(?ms)^\s*ID\s*=\s*([^\r\n]+)\s*\r?\n(.*?)(?=^\s*ID\s*=|\Z)"
    )
    updated = errors = 0
    per_language = {item["language_code"]: 0 for item in languages}
    for block in block_pattern.finditer(text):
        language_key = block.group(1).strip()[:200]
        row = connection.execute(
            "SELECT id FROM localization_strings WHERE language_key=? AND deleted_at IS NULL",
            (language_key,),
        ).fetchone()
        values = list(re.finditer(
            rf"(?ms)^\s*({label_pattern})\s*:\s*\r?\n(.*?)"
            rf"(?=^\s*(?:{label_pattern})\s*:|\Z)",
            block.group(2),
        ))
        if not row or not values:
            errors += 1
            continue
        for value in values:
            translation = _clean_ai_translation_value(value.group(2))
            if not translation:
                errors += 1
                continue
            language_code = by_label[value.group(1)]
            try:
                save_translation(
                    connection, row["id"], language_code, translation,
                    actor_id=actor_id, status="Completed",
                )
                updated += 1
                per_language[language_code] += 1
            except ValueError:
                errors += 1
    if not updated and not errors:
        raise ValueError("ID와 언어 라벨 형식의 번역 결과를 찾지 못했습니다.")
    connection.commit()
    return {"updated": updated, "errors": errors, "languages": per_language}


def detect_ai_translation_languages(connection, payload):
    """Return target languages explicitly labelled in an AI response."""
    text = str(payload or "")
    detected = set()
    for language in _active_prompt_languages(connection):
        label = re.escape(language["output_label"])
        if re.search(rf"(?m)^\s*(?:TITLE_|CONTENT_)?{label}\s*:", text):
            detected.add(language["language_code"])
    return detected


def _clean_ai_translation_value(value):
    """Remove clipboard wrappers without altering Markdown inside a translation."""
    text = re.sub(r"\n?\s*-{8,}\s*$", "", str(value or "")).strip()
    lines = text.splitlines()
    if lines and re.fullmatch(r"\s*```(?:text)?\s*", lines[0], re.IGNORECASE):
        lines.pop(0)
    if lines and re.fullmatch(r"\s*```\s*", lines[-1]):
        lines.pop()
    return "\n".join(lines).strip()


def import_ai_translation_text(connection, payload, language_code="en", actor_id=None):
    text = str(payload or "")
    if not text.strip():
        raise ValueError("AI 번역 결과를 붙여 넣어주세요.")
    if len(text) > 2_000_000:
        raise ValueError("붙여넣기 내용은 2MB 이하여야 합니다.")
    prompt_language = _prompt_language(connection, language_code)
    output_label = prompt_language["output_label"]
    accepted_labels = (
        output_label,
        f"TITLE_{output_label}",
        f"CONTENT_{output_label}",
    )
    label_pattern = "|".join(re.escape(label) for label in accepted_labels)
    # A single-language Work form may receive a combined EN/JA/YUE response.
    # Other language labels must terminate the selected value, never become part
    # of it. The combined importer remains responsible for saving those blocks.
    all_labels = []
    for language in _active_prompt_languages(connection):
        label = language["output_label"]
        all_labels.extend((label, f"TITLE_{label}", f"CONTENT_{label}"))
    all_label_pattern = "|".join(
        re.escape(label) for label in dict.fromkeys(all_labels)
    )
    pattern = re.compile(
        r"(?ms)^\s*ID\s*=\s*([^\r\n]+)\s*\r?\n(.*?)"
        r"(?=^\s*ID\s*=|\Z)"
    )
    matched, updated, errors, seen = 0, 0, 0, set()
    for match in pattern.finditer(text):
        matched += 1
        language_key = match.group(1).strip()[:200]
        values = {
            item.group(1): item.group(2).strip()
            for item in re.finditer(
                rf"(?ms)^\s*({label_pattern})\s*:\s*\r?\n(.*?)"
                rf"(?=^\s*(?:{all_label_pattern})\s*:|\Z)",
                match.group(2),
            )
        }
        translated_value = (
            values.get(output_label)
            or values.get(f"CONTENT_{output_label}")
            or values.get(f"TITLE_{output_label}")
            or ""
        )
        translation = _clean_ai_translation_value(translated_value)
        if language_key in seen or not translation:
            errors += 1; continue
        seen.add(language_key)
        row = connection.execute(
            "SELECT id FROM localization_strings WHERE language_key=? AND deleted_at IS NULL",
            (language_key,),
        ).fetchone()
        if not row:
            errors += 1; continue
        try:
            save_translation(connection, row["id"], language_code, translation,
                             actor_id=actor_id, status="Completed")
            updated += 1
        except ValueError:
            errors += 1
    if not matched:
        raise ValueError(
            f"ID=... 및 {output_label}: 형식의 번역 결과를 찾지 못했습니다."
        )
    connection.commit()
    return {"updated": updated, "errors": errors}


def repair_mixed_translation_records(connection):
    """Split legacy combined AI responses and remove stray closing fences.

    This is intentionally an explicit maintenance operation; normal requests do
    not rewrite stored content. Existing rows are only changed when a line-level
    active language label or an outer code fence is present.
    """
    languages = _active_prompt_languages(connection)
    by_label = {item["output_label"]: item["language_code"] for item in languages}
    label_pattern = "|".join(re.escape(label) for label in by_label)
    marker_pattern = re.compile(rf"(?m)^\s*({label_pattern})\s*:\s*\r?\n")
    rows = connection.execute(
        """SELECT string_id, language_code, translated_text, status
           FROM localization_translations
           WHERE translated_text IS NOT NULL"""
    ).fetchall()
    repaired_rows = split_values = 0
    for row in rows:
        original = str(row["translated_text"] or "")
        markers = list(marker_pattern.finditer(original))
        values = {}
        if markers:
            prefix = _clean_ai_translation_value(original[:markers[0].start()])
            if prefix:
                values[row["language_code"]] = prefix
            for index, marker in enumerate(markers):
                end = markers[index + 1].start() if index + 1 < len(markers) else len(original)
                value = _clean_ai_translation_value(original[marker.end():end])
                if value:
                    values[by_label[marker.group(1)]] = value
        else:
            cleaned = _clean_ai_translation_value(original)
            if cleaned != original.strip():
                values[row["language_code"]] = cleaned
        if not values:
            continue
        for language_code, value in values.items():
            existing = connection.execute(
                """SELECT translated_text FROM localization_translations
                   WHERE string_id=? AND language_code=?""",
                (row["string_id"], language_code),
            ).fetchone()
            if existing and str(existing["translated_text"] or "").strip() == value:
                continue
            save_translation(
                connection, row["string_id"], language_code, value,
                status="Completed", record_event=False,
            )
            repaired_rows += 1
            if language_code != row["language_code"]:
                split_values += 1
    connection.commit()
    return {"repaired_rows": repaired_rows, "split_values": split_values}


def qa_report(connection, language_code="en"):
    rows = connection.execute(
        """SELECT s.id, s.language_key, s.source_text, s.page_name, s.string_type,
                  t.translated_text, t.status
           FROM localization_strings s LEFT JOIN localization_translations t
             ON t.string_id=s.id AND t.language_code=? WHERE s.deleted_at IS NULL""",
        (language_code,),
    ).fetchall()
    findings = []
    for row in rows:
        source = row["source_text"] or ""
        target = row["translated_text"] or ""
        issues = []
        if not target:
            issues.append("영어 누락/빈 문자열")
        else:
            if HANGUL_RE.search(target): issues.append("한국어가 그대로 포함됨")
            if sorted(VARIABLE_RE.findall(source)) != sorted(VARIABLE_RE.findall(target)):
                issues.append("변수 누락 또는 변경")
            if len(TAG_RE.findall(source)) != len(TAG_RE.findall(target)):
                issues.append("HTML 태그 구조 변경")
            if len(MARKDOWN_RE.findall(source)) != len(MARKDOWN_RE.findall(target)):
                issues.append("Markdown 구조 변경")
            if source.count("\n") != target.count("\n"): issues.append("줄바꿈 손실")
            source_special = {c for c in source if c in "↗©®™•·→←↑↓"}
            if not source_special.issubset(set(target)): issues.append("특수문자 손실")
        for issue in issues:
            findings.append({**dict(row), "issue": issue})
    return findings


def export_rows(connection, language_code="en"):
    return [dict(row) for row in connection.execute(
        """SELECT s.id, s.language_key, s.page_name, s.component, s.string_type,
                  s.priority, s.source_text AS korean, COALESCE(t.translated_text, '') AS translation,
                  CASE WHEN s.status='Ignored' THEN 'Ignored' ELSE COALESCE(t.status, 'Pending') END AS status
           FROM localization_strings s LEFT JOIN localization_translations t
             ON t.string_id=s.id AND t.language_code=?
           WHERE s.deleted_at IS NULL ORDER BY s.id""", (language_code,)
    ).fetchall()]


def export_file(connection, language_code="en", file_type="csv"):
    rows = export_rows(connection, language_code)
    headers = ["id", "language_key", "page_name", "component", "string_type", "priority", "korean", "translation", "status"]
    def safe_cell(value):
        value = "" if value is None else str(value)
        return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value

    safe_rows = [{name: safe_cell(row[name]) for name in headers} for row in rows]
    if file_type == "xlsx":
        stream = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Localization"
        sheet.append(headers)
        for row in safe_rows: sheet.append([row[name] for name in headers])
        workbook.save(stream)
        return stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=headers)
    writer.writeheader(); writer.writerows(safe_rows)
    return ("\ufeff" + stream.getvalue()).encode("utf-8"), "text/csv; charset=utf-8"


def import_file(connection, file_storage, language_code, actor_id=None):
    filename = (file_storage.filename or "").lower()
    raw = file_storage.read(10 * 1024 * 1024 + 1)
    if len(raw) > 10 * 1024 * 1024:
        raise ValueError("파일은 10MB 이하여야 합니다.")
    if filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(raw.decode("utf-8-sig")))
        rows = list(reader)
    elif filename.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values)]
        rows = [dict(zip(headers, values_row)) for values_row in values]
    else:
        raise ValueError("CSV 또는 XLSX 파일만 가져올 수 있습니다.")
    updated = errors = 0
    for row in rows:
        try:
            string_id = int(row.get("id") or 0)
            translation = str(
                row.get("translation") or row.get(language_code)
                or row.get(language_code.upper()) or row.get("English")
                or row.get("english") or ""
            ).strip()
            status = str(row.get("status") or ("Completed" if translation else "Pending")).strip()
            if not connection.execute("SELECT 1 FROM localization_strings WHERE id=?", (string_id,)).fetchone():
                raise ValueError("unknown id")
            if translation and status == "Pending":
                status = "Completed"
            if not translation and status == "Pending":
                raise ValueError("empty pending translation")
            save_translation(connection, string_id, language_code, translation,
                             actor_id=actor_id, status=status)
            updated += 1
        except (TypeError, ValueError):
            errors += 1
    connection.commit()
    return {"updated": updated, "errors": errors}
