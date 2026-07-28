"""기존 외부자료제공 XLSX/XLSM 대장을 안전하게 일괄 등록한다."""

import io
import re
import unicodedata
from datetime import date, datetime

from openpyxl import load_workbook

from services import official_document_manager as manager
from utils import now_kst


EXPECTED_HEADERS = [
    "접수번호", "담당", "접수일", "기관명(회사명)", "사건", "요청내용",
    "특이사항", "의뢰자", "연락처", "이메일", "발송번호", "회신일",
    "위치", "영상반출여부", "반출확약서",
]
MAX_IMPORT_ROWS = 5000
MAX_IMPORT_BYTES = 20 * 1024 * 1024


class LedgerImportError(ValueError):
    pass


def _text(value):
    return str(value).strip() if value is not None else ""


def _date_text(value, required=False):
    if value in (None, ""):
        if required:
            raise LedgerImportError("필수 날짜가 비어 있습니다.")
        return None
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = _text(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    try:
        return manager.normalize_date(text, required=required)
    except manager.OfficialDocumentError:
        if required:
            raise
        return None


def _match_key(value):
    text = unicodedata.normalize("NFKC", _text(value)).replace("/", "\\")
    text = re.sub(r"\s+", " ", text).strip(" \\")
    return text.casefold()


def _folder_indexes(connection):
    rows = [
        dict(row) for row in connection.execute(
            """
            SELECT folder_name, relative_path, unc_path, folder_category, file_count
            FROM official_folder_index WHERE is_available=1
            """
        ).fetchall()
    ]
    indexes = {"name": {}, "relative": {}, "unc": {}}
    for row in rows:
        for kind, value in (
            ("name", row["folder_name"]),
            ("relative", row["relative_path"]),
            ("unc", row["unc_path"]),
        ):
            key = _match_key(value)
            if key:
                indexes[kind].setdefault(key, []).append(row)
    return indexes


def _resolve_location(raw_location, indexes):
    raw = _text(raw_location)
    if not raw:
        return None
    normalized = raw.replace("/", "\\").strip()
    candidates = []
    if normalized.startswith("\\\\"):
        candidates = indexes["unc"].get(_match_key(normalized), [])
    elif re.match(r"^[A-Za-z]:\\", normalized):
        parts = normalized.split("\\")
        for start in range(1, len(parts)):
            candidates = indexes["relative"].get(_match_key("\\".join(parts[start:])), [])
            if len(candidates) == 1:
                break
    else:
        candidates = indexes["name"].get(_match_key(normalized), [])
        if not candidates:
            candidates = indexes["relative"].get(_match_key(normalized), [])
    unique = {row["unc_path"]: row for row in candidates}
    if len(unique) != 1:
        return None
    row = next(iter(unique.values()))
    return {
        "unc_path": row["unc_path"],
        "y_path": r"Y:\07. 외부자료제공" + "\\" + row["relative_path"],
        "folder_name": row["folder_name"],
        "folder_category": row["folder_category"],
        "file_count": max(0, int(row["file_count"] or 0)),
    }


def _folder_code(location, default_code):
    label = _text(location.get("folder_category") if location else "")
    match = re.match(r"^(07[1-9])(?:\.|\s|$)", label)
    return match.group(1) if match else default_code


def import_ledger(connection, raw_bytes, filename, user, default_category, default_folder):
    if not raw_bytes:
        raise LedgerImportError("업로드할 Excel 파일을 선택해주세요.")
    if len(raw_bytes) > MAX_IMPORT_BYTES:
        raise LedgerImportError("Excel 파일은 20MB 이하여야 합니다.")
    if not _text(filename).casefold().endswith((".xlsx", ".xlsm")):
        raise LedgerImportError(".xlsx 또는 .xlsm 파일만 업로드할 수 있습니다.")
    if not manager.get_reference(connection, "category", default_category):
        raise LedgerImportError("기본 분류가 올바르지 않습니다.")
    if not manager.get_reference(connection, "folder_category", default_folder):
        raise LedgerImportError("기본 폴더구분이 올바르지 않습니다.")

    try:
        workbook = load_workbook(
            io.BytesIO(raw_bytes), read_only=True, data_only=True,
            keep_links=False, keep_vba=False,
        )
    except Exception as error:
        raise LedgerImportError("Excel 파일을 읽을 수 없습니다.") from error
    try:
        if "DB" not in workbook.sheetnames:
            raise LedgerImportError("'DB' 시트를 찾을 수 없습니다.")
        sheet = workbook["DB"]
        headers = [_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
            raise LedgerImportError("DB 시트의 열 형식이 기존 대장과 다릅니다.")
        indexes = _folder_indexes(connection)
        now_iso = now_kst().isoformat()
        result = {"inserted": 0, "skipped": 0, "linked": 0, "blank_location": 0, "errors": []}
        connection.execute("BEGIN IMMEDIATE")
        for excel_row, values in enumerate(
            sheet.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True), 2
        ):
            if excel_row > MAX_IMPORT_ROWS + 1:
                raise LedgerImportError(f"최대 {MAX_IMPORT_ROWS}행까지 업로드할 수 있습니다.")
            if not any(value not in (None, "") for value in values):
                continue
            row = dict(zip(EXPECTED_HEADERS, values))
            try:
                receipt_date = _date_text(row["접수일"], required=True)
                reply_date = _date_text(row["회신일"])
                manager_name = _text(row["담당"])
                organization = manager.canonicalize_organization(
                    connection, _text(row["기관명(회사명)"])
                )
                request_content = _text(row["요청내용"]) or _text(row["특이사항"])
                if not manager_name or not organization or not request_content:
                    raise LedgerImportError("담당, 기관명 또는 요청내용이 비어 있습니다.")
                receipt_number = _text(row["접수번호"]) or None
                duplicate = connection.execute(
                    """
                    SELECT id FROM official_documents
                    WHERE receipt_date=? AND manager=? AND
                          ((receipt_number IS NULL AND ? IS NULL) OR receipt_number=?)
                    LIMIT 1
                    """,
                    (receipt_date, manager_name, receipt_number, receipt_number),
                ).fetchone()
                if duplicate:
                    result["skipped"] += 1
                    continue

                location = _resolve_location(row["위치"], indexes)
                folder_code = _folder_code(location, default_folder)
                category = manager.get_reference(connection, "category", default_category)
                management_number = manager._next_management_number(
                    connection, receipt_date, category["label"]
                )
                video = _text(row["영상반출여부"])
                video = video if video in manager.VIDEO_EXPORT_VALUES else "해당없음"
                pledge = _text(row["반출확약서"])
                pledge = pledge if pledge in manager.EXPORT_PLEDGE_VALUES else "해당없음"
                cursor = connection.execute(
                    """
                    INSERT INTO official_documents (
                        management_number, receipt_number, manager, receipt_date, organization,
                        case_name, request_content, special_note, requester, contact, email,
                        dispatch_number, reply_date, location, normalized_unc_path,
                        video_exported, export_pledge, category_code, folder_category_code,
                        processing_result, registered_by, registered_user_id, registered_at,
                        updated_at, file_count, temp_file_status, storage_status,
                        folder_handling_type, requested_folder_path, folder_display_name,
                        folder_link_status, folder_verified_at, folder_verified_by_client
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '완료',
                              ?, ?, ?, ?, ?, 'ARCHIVED', ?, 'LINK_EXISTING', ?, ?, ?, ?, ?)
                    """,
                    (
                        management_number, receipt_number, manager_name, receipt_date, organization,
                        _text(row["사건"]) or None, request_content, _text(row["특이사항"]) or None,
                        _text(row["의뢰자"]) or None, _text(row["연락처"]) or None,
                        _text(row["이메일"]) or None, _text(row["발송번호"]) or None, reply_date,
                        location["unc_path"] if location else None,
                        location["unc_path"] if location else None,
                        video, pledge, default_category, folder_code,
                        user.get("username") or "unknown", user.get("id"), now_iso, now_iso,
                        location["file_count"] if location else 0,
                        "STORED" if location else "ARCHIVED",
                        location["y_path"] if location else None,
                        location["folder_name"] if location else None,
                        "LINK_VERIFIED" if location else None,
                        now_iso if location else None,
                        "EXCEL_IMPORT" if location else None,
                    ),
                )
                manager.log_history(
                    connection, cursor.lastrowid, "DOCUMENT_IMPORTED_FROM_EXCEL",
                    actor=user.get("username"),
                    detail={
                        "source_filename": _text(filename),
                        "source_row": excel_row,
                        "path_linked": bool(location),
                    },
                )
                result["inserted"] += 1
                if location:
                    result["linked"] += 1
                else:
                    result["blank_location"] += 1
            except Exception as error:
                result["errors"].append({"row": excel_row, "message": str(error)})
        connection.commit()
        return result
    except Exception:
        connection.rollback()
        raise
    finally:
        workbook.close()
